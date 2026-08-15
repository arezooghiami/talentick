"""
Talentick — Position Excel Service
=====================================
Import/Export گروهی پست‌های سازمانی از/به فایل Excel + دانلود قالب نمونه.

ترتیب ستون‌های فایل (import و template) — ثابت و مستند:
    1. عنوان پست *      (name)
    2. واحد سازمانی     (نام Department — باید در سازمان موجود باشد — اختیاری)
    3. سطح سازمانی       (level — عدد ۱ تا ۸، پیش‌فرض ۱)
    4. توضیحات           (description — اختیاری)

* = الزامی

پست‌ها برخلاف کاربران شناسه‌ی یکتای طبیعی (مثل موبایل) ندارند — کلید
تطبیق برای به‌روزرسانی، «عنوان پست» است (case-insensitive، در همان سازمان):
اگر پستی با همین نام از قبل وجود داشته باشد به‌روزرسانی می‌شود، وگرنه
پست جدید ساخته می‌شود. هر سطر منبع کامل مقادیر است — یعنی فیلدهای خالی
مقدار قبلی را هم بازنویسی می‌کنند (سطح خالی → ۱، واحد خالی → بدون واحد).

خروجی Export همین ستون‌ها را به‌علاوه وضعیت و تعداد کاربر دارد.
"""

from __future__ import annotations

import io
import uuid

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.organization import Department, Position
from app.models.user import User
from app.schemas.position import PositionImportResult, PositionImportRowError

# ─── ستون‌های ثابت ──────────────────────────────────────────────────────────

IMPORT_HEADERS = [
    "عنوان پست",
    "واحد سازمانی",
    "سطح سازمانی",
    "توضیحات",
]

EXPORT_HEADERS = IMPORT_HEADERS + ["وضعیت", "تعداد کاربر"]

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 26


# ─── دانلود قالب نمونه ──────────────────────────────────────────────────────

def build_template_workbook() -> bytes:
    """فایل نمونه Import — شامل ستون‌های موردنیاز + یک ردیف داده نمونه + شیت راهنما."""
    wb = Workbook()
    ws = wb.active
    ws.title = "پست‌ها"
    ws.sheet_view.rightToLeft = True
    ws.append(IMPORT_HEADERS)
    ws.append(["مدیر محصول", "فناوری اطلاعات", 3, "مسئول نقشه‌راه محصول"])
    _style_header(ws, len(IMPORT_HEADERS))

    guide = wb.create_sheet("راهنما")
    guide.sheet_view.rightToLeft = True
    guide.append(["ستون", "توضیح"])
    _style_header(guide, 2)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 70
    rows = [
        ("عنوان پست", "الزامی — حداقل ۲ حرف. اگر پستی با همین نام از قبل در سازمان وجود داشته باشد، به‌روزرسانی می‌شود."),
        ("واحد سازمانی", "نام دقیق یکی از واحدهای تعریف‌شده در سازمان (اختیاری)."),
        ("سطح سازمانی", "عددی بین ۱ (کارمند) تا ۸ (مدیرعامل) — در صورت خالی بودن، ۱ در نظر گرفته می‌شود."),
        ("توضیحات", "اختیاری."),
    ]
    for row in rows:
        guide.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Export ─────────────────────────────────────────────────────────────────

async def build_export_workbook(db: AsyncSession, positions: list[Position]) -> bytes:
    """
    فایل Excel از لیست پست‌ها (از قبل فیلترشده، در یک سازمان) می‌سازد.
    positions باید با department relationship لود شده باشند.
    """
    counts: dict[str, int] = {}
    if positions:
        org_id = positions[0].org_id
        rows = await db.execute(
            select(User.position_id, func.count())
            .where(User.org_id == org_id, User.position_id.is_not(None))
            .group_by(User.position_id)
        )
        counts = {str(pid): count for pid, count in rows.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "پست‌ها"
    ws.sheet_view.rightToLeft = True
    ws.append(EXPORT_HEADERS)
    _style_header(ws, len(EXPORT_HEADERS))

    for p in positions:
        ws.append([
            p.name,
            p.department.name if p.department else "",
            p.level,
            p.description or "",
            "فعال" if p.is_active else "غیرفعال",
            counts.get(str(p.id), 0),
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Import ─────────────────────────────────────────────────────────────────

def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


async def import_positions_from_excel(
    db: AsyncSession,
    org_id: uuid.UUID,
    file_bytes: bytes,
) -> PositionImportResult:
    """
    Import گروهی پست‌های سازمانی از فایل Excel برای یک سازمان مشخص.

    - داده‌های هر سطر اعتبارسنجی می‌شوند (عنوان، واحد، سطح).
    - «عنوان پست» کلید تطبیق است — تکراری‌های داخل خودِ فایل شناسایی و
      رد می‌شوند. پستی با همین نام در سازمان به‌روزرسانی می‌شود، وگرنه
      پست جدید ساخته می‌شود.
    - عملیات Bulk است — همه رکوردهای معتبر یک‌جا commit می‌شوند.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return PositionImportResult(
            total_rows=0, created=0, updated=0, skipped=0,
            errors=[PositionImportRowError(row=0, message=f"فایل اکسل قابل خواندن نیست: {exc}")],
        )

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    depts_result = await db.execute(select(Department).where(Department.org_id == org_id))
    depts_by_name = {d.name.strip().lower(): d for d in depts_result.scalars().all()}

    positions_result = await db.execute(select(Position).where(Position.org_id == org_id))
    positions_by_name = {p.name.strip().lower(): p for p in positions_result.scalars().all()}

    errors: list[PositionImportRowError] = []
    created = 0
    updated = 0
    skipped = 0
    seen_names: set[str] = set()

    for idx, row in enumerate(rows, start=2):
        if row is None or all(_norm(c) == "" for c in row):
            continue  # سطر کاملاً خالی — نادیده گرفته می‌شود

        name = _norm(row[0]) if len(row) > 0 else ""
        dept_name = _norm(row[1]) if len(row) > 1 else ""
        level_raw = _norm(row[2]) if len(row) > 2 else ""
        description = _norm(row[3]) if len(row) > 3 else ""

        row_error = None
        name_key = name.lower()

        if not name or len(name) < 2:
            row_error = "عنوان پست الزامی است (حداقل ۲ حرف)"
        elif name_key in seen_names:
            row_error = "عنوان پست تکراری در همین فایل"

        level = 1
        if not row_error and level_raw:
            try:
                level = int(float(level_raw))
            except ValueError:
                row_error = "سطح سازمانی باید عدد باشد"
            else:
                if not (1 <= level <= 8):
                    row_error = "سطح سازمانی باید بین ۱ تا ۸ باشد"

        dept = None
        if not row_error and dept_name:
            dept = depts_by_name.get(dept_name.lower())
            if dept is None:
                row_error = f"واحد «{dept_name}» در این سازمان یافت نشد"

        if row_error:
            errors.append(PositionImportRowError(row=idx, name=name or None, message=row_error))
            skipped += 1
            continue

        seen_names.add(name_key)
        existing = positions_by_name.get(name_key)

        if existing:
            existing.dept_id = dept.id if dept else None
            existing.level = level
            existing.description = description or None
            updated += 1
        else:
            pos = Position(
                id=uuid.uuid4(),
                org_id=org_id,
                dept_id=dept.id if dept else None,
                name=name,
                description=description or None,
                level=level,
                is_active=True,
            )
            db.add(pos)
            positions_by_name[name_key] = pos
            created += 1

    if created or updated:
        await db.commit()

    return PositionImportResult(
        total_rows=len(rows),
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
    )
