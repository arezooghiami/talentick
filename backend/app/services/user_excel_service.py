"""
Talentick — User Excel Service
=================================
Import/Export گروهی کاربران از/به فایل Excel + دانلود قالب نمونه.

ترتیب ستون‌های فایل (import و template) — ثابت و مستند:
    1. نام و نام خانوادگی *  (full_name)
    2. شماره موبایل *        (phone — شناسه‌ی اصلی یکتایی/لاگین)
    3. ایمیل                 (email — اختیاری)
    4. نقش                   (role — پیش‌فرض employee)
    5. دپارتمان               (نام Department — باید در سازمان موجود باشد)
    6. سمت                    (نام Position — باید در سازمان موجود باشد)
    7. ایمیل مدیر مستقیم      (manager_email — باید کاربر موجود باشد)

* = الزامی

خروجی Export همین ستون‌ها را به‌علاوه وضعیت/سازمان/تاریخ ثبت‌نام دارد.
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.phone import normalize_phone
from app.core.security import generate_temp_password, hash_password
from app.models.organization import Department, Position
from app.models.user import VALID_ROLES, User
from app.schemas.user import CreatedUserCredential, UserImportResult, UserImportRowError

# ─── ستون‌های ثابت ──────────────────────────────────────────────────────────

IMPORT_HEADERS = [
    "نام و نام خانوادگی",
    "شماره موبایل",
    "ایمیل",
    "نقش",
    "دپارتمان",
    "سمت",
    "ایمیل مدیر مستقیم",
]

EXPORT_HEADERS = IMPORT_HEADERS + ["وضعیت", "سازمان", "تاریخ ثبت‌نام"]

ROLE_HINT = "super_admin | org_admin | manager | employee (پیش‌فرض: employee)"

_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws, ncols: int) -> None:
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = 24


# ─── دانلود قالب نمونه ──────────────────────────────────────────────────────

def build_template_workbook() -> bytes:
    """فایل نمونه Import — شامل ستون‌های موردنیاز + یک ردیف داده نمونه + شیت راهنما."""
    wb = Workbook()
    ws = wb.active
    ws.title = "کاربران"
    ws.sheet_view.rightToLeft = True
    ws.append(IMPORT_HEADERS)
    ws.append([
        "علی رضایی", "09120000000", "ali.rezaei@example.com", "employee",
        "فناوری اطلاعات", "کارشناس Backend", "manager.email@example.com",
    ])
    _style_header(ws, len(IMPORT_HEADERS))

    guide = wb.create_sheet("راهنما")
    guide.sheet_view.rightToLeft = True
    guide.append(["ستون", "توضیح"])
    _style_header(guide, 2)
    guide.column_dimensions["A"].width = 24
    guide.column_dimensions["B"].width = 70
    rows = [
        ("نام و نام خانوادگی", "الزامی — حداقل ۲ حرف."),
        ("شماره موبایل", "الزامی و یکتا در کل پلتفرم — 09xxxxxxxxx. شناسه‌ی اصلی لاگین کاربر."),
        ("ایمیل", "اختیاری. اگر پر شود باید معتبر و یکتا باشد."),
        ("نقش", ROLE_HINT),
        ("دپارتمان", "نام دقیق یکی از دپارتمان‌های تعریف‌شده در سازمان (اختیاری)."),
        ("سمت", "نام دقیق یکی از پست‌های سازمانی تعریف‌شده (اختیاری)."),
        ("ایمیل مدیر مستقیم", "ایمیل یک کاربر موجود در همان سازمان (اختیاری)."),
    ]
    for row in rows:
        guide.append(row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Export ─────────────────────────────────────────────────────────────────

def build_export_workbook(users: list[User], org_names: dict[str, str]) -> bytes:
    """
    فایل Excel از لیست کاربران (از قبل فیلترشده) می‌سازد.
    users باید با department/position relationship لود شده باشند.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "کاربران"
    ws.sheet_view.rightToLeft = True
    ws.append(EXPORT_HEADERS)
    _style_header(ws, len(EXPORT_HEADERS))

    for u in users:
        ws.append([
            u.full_name,
            u.phone,
            u.email or "",
            u.role,
            u.department.name if u.department else "",
            u.position.name if u.position else "",
            "",  # ایمیل مدیر — در export نمایش داده نمی‌شود مگر جدا resolve شود
            "فعال" if u.is_active else "غیرفعال",
            org_names.get(str(u.org_id), ""),
            u.created_at.strftime("%Y-%m-%d") if u.created_at else "",
        ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Import ─────────────────────────────────────────────────────────────────

def _norm(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


async def import_users_from_excel(
    db: AsyncSession,
    org_id: uuid.UUID,
    file_bytes: bytes,
    *,
    update_existing: bool = False,
) -> UserImportResult:
    """
    Import گروهی کاربران از فایل Excel برای یک سازمان مشخص.

    - داده‌های هر سطر اعتبارسنجی می‌شوند (موبایل، ایمیل، نقش، دپارتمان/سمت/مدیر).
    - شماره موبایل شناسه‌ی اصلی یکتایی است — تکراری‌های داخل خودِ فایل
      شناسایی و رد می‌شوند.
    - اگر update_existing=True باشد، کاربران با موبایل تکراری (نسبت به DB)
      به‌روزرسانی می‌شوند؛ در غیر این صورت آن سطر رد (skip) می‌شود.
    - عملیات Bulk است — همه رکوردهای معتبر یک‌جا commit می‌شوند.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        return UserImportResult(
            total_rows=0, created=0, updated=0, skipped=0,
            errors=[UserImportRowError(row=0, message=f"فایل اکسل قابل خواندن نیست: {exc}")],
        )

    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # ─── پیش‌بارگذاری داده‌های سازمان برای جلوگیری از N+1 query ──────────
    depts_result = await db.execute(select(Department).where(Department.org_id == org_id))
    depts_by_name = {d.name.strip().lower(): d for d in depts_result.scalars().all()}

    positions_result = await db.execute(select(Position).where(Position.org_id == org_id))
    positions_by_name = {p.name.strip().lower(): p for p in positions_result.scalars().all()}

    org_users_result = await db.execute(select(User).where(User.org_id == org_id))
    org_users_all = org_users_result.scalars().all()
    users_by_email_in_org = {u.email.lower(): u for u in org_users_all if u.email}

    all_users_result = await db.execute(select(User))
    all_users_all = all_users_result.scalars().all()
    all_users_by_phone = {u.phone: u for u in all_users_all}

    errors: list[UserImportRowError] = []
    created = 0
    updated = 0
    skipped = 0
    seen_phones: set[str] = set()

    # سطرهایی که باید ساخته/به‌روزرسانی شوند را جمع می‌کنیم تا در آخر یک‌جا commit شود
    to_create: list[User] = []
    created_credentials: list[CreatedUserCredential] = []

    for idx, row in enumerate(rows, start=2):
        if row is None or all(_norm(c) == "" for c in row):
            continue  # سطر کاملاً خالی — نادیده گرفته می‌شود

        full_name = _norm(row[0]) if len(row) > 0 else ""
        phone_raw = _norm(row[1]) if len(row) > 1 else ""
        email = _norm(row[2]).lower() if len(row) > 2 else ""
        role = _norm(row[3]).lower() if len(row) > 3 and _norm(row[3]) else "employee"
        dept_name = _norm(row[4]) if len(row) > 4 else ""
        position_name = _norm(row[5]) if len(row) > 5 else ""
        manager_email = _norm(row[6]).lower() if len(row) > 6 else ""

        row_error = None
        phone = ""

        if not full_name or len(full_name) < 2:
            row_error = "نام و نام خانوادگی الزامی است (حداقل ۲ حرف)"
        else:
            try:
                phone = normalize_phone(phone_raw)
            except ValueError:
                row_error = "شماره موبایل الزامی و باید معتبر باشد (09xxxxxxxxx)"

        if not row_error and email and ("@" not in email or "." not in email.split("@")[-1]):
            row_error = "ایمیل نامعتبر است"
        elif not row_error and role not in VALID_ROLES:
            row_error = f"نقش نامعتبر است — مقادیر مجاز: {', '.join(sorted(VALID_ROLES))}"
        elif not row_error and phone in seen_phones:
            row_error = "شماره موبایل تکراری در همین فایل"

        dept = None
        if not row_error and dept_name:
            dept = depts_by_name.get(dept_name.lower())
            if dept is None:
                row_error = f"دپارتمان «{dept_name}» در این سازمان یافت نشد"

        position = None
        if not row_error and position_name:
            position = positions_by_name.get(position_name.lower())
            if position is None:
                row_error = f"سمت «{position_name}» در این سازمان یافت نشد"

        manager = None
        if not row_error and manager_email:
            manager = users_by_email_in_org.get(manager_email)
            if manager is None:
                row_error = f"مدیر مستقیم با ایمیل «{manager_email}» در این سازمان یافت نشد"

        if row_error:
            errors.append(UserImportRowError(row=idx, phone=phone or None, email=email or None, message=row_error))
            skipped += 1
            continue

        seen_phones.add(phone)
        existing = all_users_by_phone.get(phone)

        if existing:
            if not update_existing:
                errors.append(UserImportRowError(
                    row=idx, phone=phone, email=email or None,
                    message="این شماره موبایل قبلاً در سیستم ثبت شده — برای به‌روزرسانی، update_existing را فعال کنید",
                ))
                skipped += 1
                continue
            if str(existing.org_id) != str(org_id):
                errors.append(UserImportRowError(
                    row=idx, phone=phone, email=email or None,
                    message="این شماره موبایل متعلق به کاربری در سازمان دیگری است — قابل به‌روزرسانی نیست",
                ))
                skipped += 1
                continue

            existing.full_name = full_name
            existing.role = role
            existing.dept_id = dept.id if dept else None
            existing.position_id = position.id if position else None
            existing.manager_id = manager.id if manager else None
            if email:
                existing.email = email
            updated += 1
        else:
            temp_password = generate_temp_password()
            new_user = User(
                id=uuid.uuid4(),
                org_id=org_id,
                email=email or None,
                full_name=full_name,
                hashed_password=hash_password(temp_password),
                role=role,
                dept_id=dept.id if dept else None,
                position_id=position.id if position else None,
                manager_id=manager.id if manager else None,
                phone=phone,
                is_active=True,
                is_email_verified=False,
                # رمز موقت را ادمین از طریق created_users در همین پاسخ
                # می‌بیند و باید دستی به کاربر بدهد — کاربر موظف به تغییر
                # آن در اولین ورود است (بدون سرویس ایمیل/پیامک در این جریان).
                must_change_password=True,
            )
            db.add(new_user)
            to_create.append(new_user)
            created_credentials.append(
                CreatedUserCredential(phone=phone, email=email or None, temp_password=temp_password)
            )
            # جلوگیری از تکراری و امکان ارجاع «مدیر مستقیم» به کاربر تازه‌ساخته‌شده در همین فایل
            all_users_by_phone[phone] = new_user
            if email:
                users_by_email_in_org[email] = new_user
            created += 1

    if created or updated:
        await db.commit()

    return UserImportResult(
        total_rows=len(rows),
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
        created_users=created_credentials,
    )
